from collections import Counter
from itertools import combinations
from pathlib import Path

import pandas as pd


def calcular_resultado(dados_mantidos, faces):
    acertos = 0
    desafios = 0
    adaptacoes = 0

    for dado in dados_mantidos:
        resultado_face = faces[dado]
        acertos += resultado_face[0]
        desafios += resultado_face[1]
        adaptacoes += resultado_face[2]

    return acertos, desafios, adaptacoes


def gerar_combinacoes_mantidas(rolagem, quantidade_mantida):
    return combinations(rolagem, quantidade_mantida)


def registrar_resultado(contador, resultado):
    contador[resultado] += 1


def criar_tabela_resultados(contador, total_ocorrencias):
    linhas = []

    for resultado, ocorrencias in contador.items():
        acertos, desafios, adaptacoes = resultado
        probabilidade_percentual = ocorrencias / total_ocorrencias * 100

        linhas.append({
            "Acertos": acertos,
            "Desafios": desafios,
            "Adaptacoes": adaptacoes,
            "Ocorrencias": ocorrencias,
            "Probabilidade_percentual": probabilidade_percentual
        })

    tabela = pd.DataFrame(linhas)

    if tabela.empty:
        return tabela

    tabela = tabela.sort_values(
        by=["Acertos", "Desafios", "Adaptacoes"]
    ).reset_index(drop=True)

    return tabela


def criar_contador():
    return Counter()


def ajustar_largura_colunas(planilha):
    for coluna in planilha.columns:
        maior_tamanho = 0
        letra_coluna = coluna[0].column_letter

        for celula in coluna:
            valor = "" if celula.value is None else str(celula.value)
            maior_tamanho = max(maior_tamanho, len(valor))

        planilha.column_dimensions[letra_coluna].width = min(
            maior_tamanho + 2,
            40
        )


def formatar_planilha(planilha):
    from openpyxl.styles import Alignment, Font, PatternFill

    cor_cabecalho = "1F4E78"
    preenchimento = PatternFill(
        fill_type="solid",
        fgColor=cor_cabecalho
    )

    for celula in planilha[1]:
        celula.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celula.fill = preenchimento
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    for linha in planilha.iter_rows():
        for celula in linha:
            celula.alignment = Alignment(
                vertical="center"
            )

    ajustar_largura_colunas(planilha)


def salvar_excel_completo(
    tabelas,
    caminho_arquivo,
    colunas_percentuais=None
):
    caminho = Path(caminho_arquivo)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    if colunas_percentuais is None:
        colunas_percentuais = {}

    with pd.ExcelWriter(
        caminho,
        engine="openpyxl"
    ) as writer:
        for nome_aba, tabela in tabelas.items():
            tabela.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False
            )

            planilha = writer.sheets[nome_aba]
            formatar_planilha(planilha)

            for nome_coluna in colunas_percentuais.get(nome_aba, []):
                if nome_coluna not in tabela.columns:
                    continue

                indice_coluna = list(tabela.columns).index(nome_coluna) + 1

                for linha in range(2, len(tabela) + 2):
                    planilha.cell(
                        row=linha,
                        column=indice_coluna
                    ).number_format = '0.0000"%"'

    return caminho
